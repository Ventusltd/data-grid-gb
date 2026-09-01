"""Parse ETYS Appendix B and D into one clean network model.

WHAT COMES OUT
--------------
derived/gb-transmission-network.v1.json

  sites      every substation the system operator names, with its code,
             its transmission owner and the voltages present there
  nodes      the busbar-level identities that circuits actually connect,
             each carrying its site, its voltage and its fault level where
             the operator publishes one
  circuits   node to node, with length, type, R / X / B on a 100 MVA base
             and seasonal ratings
  transformers  node to node, with impedance and rating
  changes    the circuit and transformer changes the operator has already
             published for 2026/27 to 2033/34, which is the only public
             statement of what the network is about to become
  compensation  reactive plant, because a connection's reactive
             requirement is answered by what is already installed
  interconnectors

WHAT IS NOT DONE HERE
---------------------
No power flow is solved and none is implied. This is the published model's
parameters, restated in a form software can read. A rating is a rating and
a fault level is a fault level; neither is a statement about whether any
particular project can connect anywhere, which depends on queue position,
committed connections, consent and commercial terms that no published
appendix contains.

NODE CODES
----------
A node is a site code of up to four characters, then a digit for the
voltage level, then a suffix identifying the busbar or bay. The digit
convention is not documented in the appendix, so it is DERIVED here by
counting how each digit co-occurs with the voltages its site declares, and
the counts are published in the product rather than asserted. Observed on
the 2025 edition: 1 -> 132 kV, 2 -> 275 kV, 4 -> 400 kV, and a node whose
inferred voltage is not among its site's declared voltages is flagged
rather than silently corrected.

    python pipelines/build_network_model.py
"""

import io
import json
import os
import re
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
SOURCES = os.path.join(REPO, "sources")
DERIVED = os.path.join(REPO, "derived")

EDITION = "2025"
OWNERS = {"a": "SHET", "b": "SPT", "c": "NGET", "d": "OFTO"}
VOLTAGE_DIGIT = {"1": 132, "2": 275, "4": 400}
FAULT_HEADER_KEYS = {
    "Three Phase Initial Peak Current (kA)": "three_phase_initial_peak_current_ka",
    "Three Phase RMS Break Current (kA)": "three_phase_rms_break_current_ka",
    "Symmetrical Three Phase RMS Break Current (kA)": "three_phase_rms_break_current_ka",
    "Three Phase DC Break Current (kA)": "three_phase_dc_break_current_ka",
    "Three Phase Peak Break Current (kA)": "three_phase_peak_break_current_ka",
    "Asymmetrical Three Phase Peak Break Current (kA)": "three_phase_peak_break_current_ka",
    "Single Phase Initial Peak Current (kA)": "single_phase_initial_peak_current_ka",
    "Single Phase RMS Break Current (kA)": "single_phase_rms_break_current_ka",
    "Symmetrical Single Phase RMS Break Current (kA)": "single_phase_rms_break_current_ka",
    "Single Phase DC Break Current (kA)": "single_phase_dc_break_current_ka",
    "Single Phase Peak Break Current (kA)": "single_phase_peak_break_current_ka",
    "Asymmetrical Single Phase Peak Break Current (kA)": "single_phase_peak_break_current_ka",
}
FAULT_KEYS = (
    "three_phase_initial_peak_current_ka",
    "three_phase_rms_break_current_ka",
    "three_phase_dc_break_current_ka",
    "three_phase_peak_break_current_ka",
    "single_phase_initial_peak_current_ka",
    "single_phase_rms_break_current_ka",
    "single_phase_dc_break_current_ka",
    "single_phase_peak_break_current_ka",
)


def cells(worksheet, min_row=1):
    for row in worksheet.iter_rows(min_row=min_row, values_only=True):
        yield row


def is_header(row):
    return any(isinstance(v, str) and v.strip() in ("Node 1", "Node1", "Site Code",
                                                    "Site Name", "OFTO", "Wind Farm",
                                                    "Interconnector Name")
               for v in row if v is not None)


def rows_after_header(worksheet):
    """Yield data rows, skipping the title block and the header line.

    The appendix puts a title in the first rows and the header wherever it
    lands, so the header is found rather than assumed at a fixed offset.
    """
    started = False
    for row in cells(worksheet):
        if not started:
            if is_header(row):
                started = True
            continue
        if all(v in (None, "") for v in row):
            continue
        yield row


def number(value):
    if value in (None, ""):
        return None
    try:
        return round(float(value), 6)
    except (TypeError, ValueError):
        return None


def text(value):
    return None if value in (None, "") else str(value).strip()


def main():
    import openpyxl

    book = openpyxl.load_workbook(
        os.path.join(SOURCES, f"etys-{EDITION}-appendix-b-system-technical-data.xlsx"),
        read_only=True, data_only=True)

    # ── sites ────────────────────────────────────────────────────────────
    sites = {}
    for suffix, owner in OWNERS.items():
        for row in rows_after_header(book[f"B-1-1{suffix}"]):
            values = [v for v in row if v not in (None, "")]
            if len(values) < 3:
                continue
            code, name, kv = text(values[0]), text(values[1]), number(values[-1])
            if not code or not name or kv is None or len(code) > 6:
                continue
            site = sites.setdefault(code.upper(), {
                "code": code.upper(), "name": name,
                "transmission_owner": owner, "voltages_kv": []})
            if int(kv) not in site["voltages_kv"]:
                site["voltages_kv"].append(int(kv))
    for site in sites.values():
        site["voltages_kv"].sort(reverse=True)

    # ── circuits, transformers, changes, compensation ────────────────────
    def branch_rows(sheet, kind, with_year=False, lead_labels=0):
        out = []
        for row in rows_after_header(book[sheet]):
            values = list(row)
            if lead_labels:
                labels = [text(v) for v in values[:lead_labels]]
                values = values[lead_labels:]
            else:
                labels = []
            node1, node2 = text(values[0]), text(values[1])
            if not node1 or not node2:
                continue
            rest = values[2:]
            record = {"node_1": node1.upper(), "node_2": node2.upper()}
            if labels:
                record["labels"] = [l for l in labels if l]
            if with_year:
                record["year"] = text(rest[0])
                record["status"] = text(rest[1])
                rest = rest[2:]
            if kind == "circuit":
                keys = ["ohl_km", "cable_km", "circuit_type", "r_pct_100mva",
                        "x_pct_100mva", "b_pct_100mva", "winter_mva",
                        "spring_mva", "summer_mva", "autumn_mva"]
            else:
                keys = ["r_pct_100mva", "x_pct_100mva", "b_pct_100mva",
                        "rating_mva", "voltage_ratio_kv"]
            for index, key in enumerate(keys):
                if index >= len(rest):
                    break
                value = rest[index]
                record[key] = text(value) if key in ("circuit_type", "voltage_ratio_kv") \
                    else number(value)
            out.append(record)
        return out

    circuits, transformers, changes, compensation = [], [], [], []
    for suffix, owner in OWNERS.items():
        lead = 2 if suffix == "d" else 0
        for record in branch_rows(f"B-2-1{suffix}", "circuit", lead_labels=lead):
            record["transmission_owner"] = owner
            circuits.append(record)
        lead_change = 1 if suffix == "d" else 0
        for record in branch_rows(f"B-2-2{suffix}", "circuit", with_year=True,
                                  lead_labels=lead_change):
            record["transmission_owner"] = owner
            record["asset"] = "circuit"
            changes.append(record)
        lead_tx = 2 if suffix == "d" else 0
        for record in branch_rows(f"B-3-1{suffix}", "transformer", lead_labels=lead_tx):
            record["transmission_owner"] = owner
            transformers.append(record)
        lead_txc = 1 if suffix == "d" else 0
        for record in branch_rows(f"B-3-2{suffix}", "transformer", with_year=True,
                                  lead_labels=lead_txc):
            record["transmission_owner"] = owner
            record["asset"] = "transformer"
            changes.append(record)
        for row in rows_after_header(book[f"B-4-1{suffix}"]):
            values = [v for v in row]
            if suffix == "d":
                values = values[1:]
            name, node = text(values[0]), text(values[1])
            if not node:
                continue
            compensation.append({
                "transmission_owner": owner, "site_name": name,
                "node": node.upper(), "unit": text(values[2]),
                "mvar_generation": number(values[3]),
                "mvar_absorption": number(values[4]),
                "type": text(values[5]),
                "connection_kv": number(values[6]) if len(values) > 6 else None})

    interconnectors = []
    for row in rows_after_header(book["B-5-1"]):
        values = list(row)
        name = text(values[0])
        if not name:
            continue
        interconnectors.append({
            "name": name, "existing": text(values[1]),
            "planned_from_year": text(values[2]),
            "node_1": text(values[3]), "node_2": text(values[4]),
            "type": text(values[5]), "rated_kv": number(values[6]),
            "length_km": number(values[7])})
    book.close()

    # ── nodes, and the derived voltage-digit convention ──────────────────
    digit_counts = defaultdict(Counter)
    node_names = set()
    for record in circuits + transformers + changes:
        node_names.update([record["node_1"], record["node_2"]])
    for node in node_names:
        code, digit = node[:4], (node[4] if len(node) > 4 else "")
        if digit.isdigit() and code in sites:
            for kv in sites[code]["voltages_kv"]:
                digit_counts[digit][kv] += 1

    nodes, unresolved = {}, 0
    for node in sorted(node_names):
        code, digit = node[:4], (node[4] if len(node) > 4 else "")
        site = sites.get(code)
        kv = VOLTAGE_DIGIT.get(digit)
        consistent = bool(site and kv and kv in site["voltages_kv"])
        if not consistent:
            unresolved += 1
        nodes[node] = {
            "node": node, "site_code": code,
            "site_name": site["name"] if site else None,
            "transmission_owner": site["transmission_owner"] if site else None,
            "voltage_kv": kv,
            # Named, not hidden: a node whose inferred voltage is not one the
            # site declares is a fact about the convention, not a defect to
            # paper over.
            "voltage_consistent_with_site": consistent}

    # ── fault levels ─────────────────────────────────────────────────────
    fault_scenarios = []
    for label, filename in (("peak", f"etys-{EDITION}-appendix-d-fault-levels-peak.xlsx"),
                            ("minimum", f"etys-{EDITION}-appendix-d-fault-levels-minimum.xlsx")):
        path = os.path.join(SOURCES, filename)
        if not os.path.exists(path):
            continue
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in book.sheetnames:
            if not re.fullmatch(r"D[123]\.\d", sheet_name):
                continue
            sheet = book[sheet_name]
            title = text(sheet.cell(1, 1).value) or ""
            header = tuple(text(sheet.cell(2, column).value) for column in range(1, 11))
            metric_keys = tuple(FAULT_HEADER_KEYS.get(value) for value in header[2:])
            if (header[:2] != ("Location", "Voltage (kV)")
                    or None in metric_keys or metric_keys != FAULT_KEYS):
                raise ValueError(f"{sheet_name}: unknown Appendix D schema: {header}")
            year_index = re.search(r"\b(?:Yr|Year)\s*(\d+)\b", title, re.I)
            winter = re.search(r"\b(\d{4}/\d{2})\b", title)
            if not year_index or not winter:
                raise ValueError(f"{sheet_name}: scenario missing from title: {title!r}")
            owner = {"1": "SHET", "2": "SPT", "3": "NGET"}[sheet_name[1]]
            owner_sites = {code: site for code, site in sites.items()
                           if site["transmission_owner"] == owner}
            sites_by_name = {site["name"].upper(): code for code, site in owner_sites.items()}
            for source_row, row in enumerate(
                    sheet.iter_rows(min_row=3, max_col=10, values_only=True), start=3):
                if not isinstance(row[0], str) or not isinstance(row[1], (int, float)):
                    continue
                location = " ".join(row[0].split()).upper()
                prefix = location[:4]
                site_code = prefix if prefix in owner_sites else sites_by_name.get(location)
                record = {
                    "demand_case": label, "transmission_owner": owner,
                    "scenario_year_index": int(year_index.group(1)),
                    "winter": winter.group(1), "location": location,
                    "site_code": site_code, "voltage_kv": float(row[1]),
                    "source_sheet": sheet_name, "source_row": source_row,
                    "published_metric_labels": list(header[2:]),
                }
                for key, value in zip(metric_keys, row[2:]):
                    if not isinstance(value, (int, float)):
                        raise ValueError(f"{sheet_name}/{source_row}: nonnumeric {key}")
                    record[key] = float(value)
                fault_scenarios.append(record)
        book.close()
    fault_scenarios.sort(key=lambda row: (
        row["demand_case"], row["transmission_owner"], row["winter"],
        row["location"], row["voltage_kv"], row["source_sheet"], row["source_row"]))

    product = {
        "schema": "data-grid-gb.transmission-network.v1",
        "source": {
            "publisher": "NESO",
            "publication": f"Electricity Ten Year Statement {EDITION}",
            "appendices": ["B - system technical data",
                           "D - fault levels (peak and minimum)"],
            "note": "parameters as published; no power flow is solved here",
        },
        "not_a_connection_assessment": (
            "A rating is a rating and a fault level is a fault level. Neither "
            "states whether any project can connect at a node, which depends "
            "on queue position, committed connections, consent and commercial "
            "terms that no published appendix contains."),
        "node_code_convention": {
            "form": "site code (up to 4 characters) + voltage digit + busbar or bay suffix",
            "voltage_digit": VOLTAGE_DIGIT,
            "derived_not_documented": True,
            "observed_digit_to_site_voltage_counts": {
                digit: dict(counter.most_common(4))
                for digit, counter in sorted(digit_counts.items())},
            "nodes_whose_voltage_is_not_declared_by_their_site": unresolved,
        },
        "counts": {
            "sites": len(sites), "nodes": len(nodes), "circuits": len(circuits),
            "transformers": len(transformers), "planned_changes": len(changes),
            "reactive_compensation_units": len(compensation),
            "interconnectors": len(interconnectors),
            "published_fault_current_scenarios": len(fault_scenarios),
            "fault_current_scenarios_with_site_code":
                sum(1 for row in fault_scenarios if row["site_code"]),
        },
        "sites": sorted(sites.values(), key=lambda s: s["code"]),
        "nodes": [nodes[k] for k in sorted(nodes)],
        "circuits": circuits,
        "transformers": transformers,
        "planned_changes": changes,
        "reactive_compensation": compensation,
        "interconnectors": interconnectors,
        "fault_current_metrics": list(FAULT_KEYS),
        "fault_current_scenarios": fault_scenarios,
    }

    os.makedirs(DERIVED, exist_ok=True)
    out = os.path.join(DERIVED, "gb-transmission-network.v1.json")
    io.open(out, "w", encoding="utf-8", newline="\n").write(
        json.dumps(product, ensure_ascii=False, indent=1) + "\n")
    size = os.path.getsize(out) / 1024 / 1024
    print(f"wrote derived/gb-transmission-network.v1.json ({size:.1f} MB)")
    for key, value in product["counts"].items():
        print(f"  {key:<44} {value:>7,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
