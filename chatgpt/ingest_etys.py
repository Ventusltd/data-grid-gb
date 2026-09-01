#!/usr/bin/env python3
"""Deterministically normalise pinned NESO ETYS 2025 workbooks."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Iterable

import openpyxl

ROOT = Path(__file__).resolve().parent
TO_BY_SUFFIX = {"a": "SHET", "b": "SPT", "c": "NGET", "d": "OFTO"}
FAULT_COLUMNS = (
    "three_phase_initial_peak_current_ka",
    "three_phase_rms_break_current_ka",
    "three_phase_dc_break_current_ka",
    "three_phase_peak_break_current_ka",
    "single_phase_initial_peak_current_ka",
    "single_phase_rms_break_current_ka",
    "single_phase_dc_break_current_ka",
    "single_phase_peak_break_current_ka",
)
FAULT_HEADER_KEYS = {
    "Three Phase Initial Peak Current (kA)": "three_phase_initial_peak_current_ka",
    "Three Phase RMS Break Current (kA)": "three_phase_rms_break_current_ka",
    "Symmetrical Three Phase RMS Break Current (kA)": "three_phase_rms_break_current_ka",
    "Three Phase DC Break Current (kA)": "three_phase_dc_break_current_ka",
    "Three Phase Peak Break Current (kA)": "three_phase_peak_break_current_ka",
    "Asymmetrical Three Phase Peak Break Current (kA)": "three_phase_peak_break_current_ka",
    "Single Phase Initial Peak Current (kA)": "single_phase_initial_peak_current_ka",
    "Single Phase RMS Break Current (kA)": "single_phase_rms_break_current_ka",
    "Symmetrical Single Phase RMS Break Current (kA)": "single_phase_rms_break_current_ka",
    "Single Phase DC Break Current (kA)": "single_phase_dc_break_current_ka",
    "Single Phase Peak Break Current (kA)": "single_phase_peak_break_current_ka",
    "Asymmetrical Single Phase Peak Break Current (kA)": "single_phase_peak_break_current_ka",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def pinned(path: Path, artifact: str, ledger: dict[str, Any]) -> dict[str, Any]:
    spec = ledger["artifacts"][artifact]
    actual = {"bytes": path.stat().st_size, "sha256": sha256(path)}
    if actual["bytes"] != spec["bytes"] or actual["sha256"] != spec["sha256"]:
        raise ValueError(f"{artifact}: source bytes do not match pinned ledger: {actual}")
    return {"artifact": artifact, "url": spec["url"], **actual}


def clean(value: Any) -> Any:
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def records(ws: Any, required: tuple[str, ...]) -> Iterable[dict[str, Any]]:
    header = None
    for row in ws.iter_rows(values_only=True):
        values = tuple(clean(v) for v in row)
        if header is None:
            if all(name in values for name in required):
                header = values
            continue
        if not any(v not in (None, "") for v in values):
            continue
        yield {str(key): value for key, value in zip(header, values) if key not in (None, "")}
    if header is None:
        raise ValueError(f"{ws.title}: required header not found: {required}")


def sheet_to(ws: Any) -> str:
    match = re.fullmatch(r"B-[1234]-[12]([abcd])", ws.title)
    if not match:
        raise ValueError(f"cannot derive transmission owner from {ws.title}")
    return TO_BY_SUFFIX[match.group(1)]


def has_header(ws: Any, required: tuple[str, ...]) -> bool:
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        values = tuple(clean(v) for v in row)
        if all(name in values for name in required):
            return True
    return False


def parse_sites(wb: Any) -> list[dict[str, Any]]:
    output = []
    for suffix, owner in TO_BY_SUFFIX.items():
        ws = wb[f"B-1-1{suffix}"]
        for row in records(ws, ("Site Code", "Site Name", "Voltage (kV)")):
            code, name, kv = row["Site Code"], row["Site Name"], row["Voltage (kV)"]
            if not isinstance(code, str) or not isinstance(name, str) or not isinstance(kv, (int, float)):
                continue
            output.append({"owner": owner, "site_code": code.strip().upper(),
                           "site_name": name.strip(), "voltage_kv": float(kv)})
    return sorted(output, key=lambda x: (x["owner"], x["site_code"], x["voltage_kv"], x["site_name"]))


def parse_equipment(wb: Any, prefix: str, kind: str, changed: bool) -> list[dict[str, Any]]:
    output = []
    for suffix in TO_BY_SUFFIX:
        ws = wb[f"{prefix}{suffix}"]
        if has_header(ws, ("Node1", "Node2")):
            iterator = records(ws, ("Node1", "Node2"))
            n1, n2 = "Node1", "Node2"
        else:
            iterator = records(ws, ("Node 1", "Node 2"))
            n1, n2 = "Node 1", "Node 2"
        for row in iterator:
            a, b = row.get(n1), row.get(n2)
            if not isinstance(a, str) or not isinstance(b, str):
                continue
            item = {"owner": TO_BY_SUFFIX[suffix], "kind": kind,
                    "node_1": a.strip().upper(), "node_2": b.strip().upper(),
                    "source_sheet": ws.title}
            for source, target in (
                ("Year", "year"), ("Status", "status"),
                ("OHL Length (km)", "ohl_length_km"),
                ("Cable Length (km)", "cable_length_km"),
                ("Circuit Type", "circuit_type"),
                ("R (% on 100 MVA)", "r_percent_100mva"),
                ("R (% on 100MVA)", "r_percent_100mva"),
                ("X (% on 100 MVA)", "x_percent_100mva"),
                ("X (% on 100MVA)", "x_percent_100mva"),
                ("B (% on 100 MVA)", "b_percent_100mva"),
                ("B (% on 100MVA)", "b_percent_100mva"),
                ("Rating (MVA)", "rating_mva"),
                ("Winter Rating (MVA)", "winter_rating_mva"),
                ("Spring Rating (MVA)", "spring_rating_mva"),
                ("Summer Rating (MVA)", "summer_rating_mva"),
                ("Autumn Rating (MVA)", "autumn_rating_mva"),
            ):
                if source in row and row[source] not in (None, ""):
                    item[target] = clean(row[source])
            if changed and ("year" not in item or "status" not in item):
                raise ValueError(f"{ws.title}: planned equipment lacks Year/Status")
            output.append(item)
    return sorted(output, key=lambda x: (x["owner"], x["node_1"], x["node_2"], str(x.get("year", ""))))


def parse_faults(path: Path, demand_case: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    output = []
    for ws in wb.worksheets:
        if not re.fullmatch(r"D[123]\.\d", ws.title):
            continue
        title = clean(ws.cell(1, 1).value)
        header = tuple(clean(ws.cell(2, col).value) for col in range(1, 11))
        metric_keys = tuple(FAULT_HEADER_KEYS.get(name) for name in header[2:])
        if (header[:2] != ("Location", "Voltage (kV)") or None in metric_keys
                or metric_keys != FAULT_COLUMNS):
            raise ValueError(f"{ws.title}: fault schema drift: {header}")
        year_index = re.search(r"\b(?:Yr|Year)\s*(\d+)\b", str(title), re.I)
        winter = re.search(r"\b(\d{4}/\d{2})\b", str(title))
        if not year_index or not winter:
            raise ValueError(f"{ws.title}: scenario not found in {title!r}")
        owner = {"1": "SHET", "2": "SPT", "3": "NGET"}[ws.title[1]]
        for source_row, row in enumerate(
                ws.iter_rows(min_row=3, max_col=10, values_only=True), start=3):
            if not isinstance(row[0], str) or not isinstance(row[1], (int, float)):
                continue
            item = {"demand_case": demand_case, "owner": owner,
                    "scenario_year_index": int(year_index.group(1)),
                    "winter": winter.group(1), "node": clean(row[0]).upper(),
                    "voltage_kv": float(row[1]), "source_sheet": ws.title,
                    "source_row": source_row}
            item["published_metric_labels"] = list(header[2:])
            for key, value in zip(metric_keys, row[2:]):
                if not isinstance(value, (int, float)):
                    raise ValueError(f"{ws.title}/{row[0]}: nonnumeric {key}")
                item[key] = float(value)
            output.append(item)
    wb.close()
    return sorted(output, key=lambda x: (x["demand_case"], x["owner"], x["winter"], x["node"]))


def unique(items: list[dict[str, Any]], fields: tuple[str, ...], label: str) -> None:
    seen = set()
    for item in items:
        key = tuple(item.get(field) for field in fields)
        if key in seen:
            raise ValueError(f"duplicate {label}: {key}")
        seen.add(key)


def build(args: argparse.Namespace) -> dict[str, Any]:
    ledger = json.loads((ROOT / "sources.json").read_text(encoding="utf-8"))
    paths = {"appendix_b": args.appendix_b, "fault_peak": args.fault_peak,
             "fault_minimum": args.fault_minimum}
    provenance = {key: pinned(path, key, ledger) for key, path in paths.items()}
    wb = openpyxl.load_workbook(args.appendix_b, read_only=True, data_only=True)
    sites = parse_sites(wb)
    circuits = parse_equipment(wb, "B-2-1", "circuit", False)
    circuit_changes = parse_equipment(wb, "B-2-2", "circuit_change", True)
    transformers = parse_equipment(wb, "B-3-1", "transformer", False)
    transformer_changes = parse_equipment(wb, "B-3-2", "transformer_change", True)
    wb.close()
    faults = parse_faults(args.fault_peak, "peak") + parse_faults(args.fault_minimum, "minimum")
    unique(sites, ("owner", "site_code", "voltage_kv"), "site voltage")
    unique(faults, ("demand_case", "owner", "source_sheet", "source_row"),
           "fault source row")
    return {
        "schema": "data-grid-gb.etys.normalized.v1",
        "edition": "ETYS 2025",
        "claim_boundary": "topology_and_equipment_parameters_not_a_solved_power_flow_case",
        "provenance": provenance,
        "counts": {"sites": len(sites), "circuits": len(circuits),
                   "circuit_changes": len(circuit_changes), "transformers": len(transformers),
                   "transformer_changes": len(transformer_changes), "fault_scenarios": len(faults)},
        "sites": sites, "circuits": circuits, "circuit_changes": circuit_changes,
        "transformers": transformers, "transformer_changes": transformer_changes,
        "fault_scenarios": faults,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--appendix-b", type=Path, required=True)
    parser.add_argument("--fault-peak", type=Path, required=True)
    parser.add_argument("--fault-minimum", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    product = build(args)
    payload = (json.dumps(product, ensure_ascii=False, sort_keys=True,
                          separators=(",", ":")) + "\n").encode("utf-8")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_bytes(payload)
    args.output.with_suffix(args.output.suffix + ".sha256").write_text(
        hashlib.sha256(payload).hexdigest() + "  " + args.output.name + "\n", encoding="ascii", newline="\n")
    print(json.dumps({"output": str(args.output), "sha256": hashlib.sha256(payload).hexdigest(),
                      "counts": product["counts"]}, indent=2))


if __name__ == "__main__":
    main()
